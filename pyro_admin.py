import asyncio

import openpyxl
from pyrogram import Client, filters
from pyrogram.errors import FloodWait
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton

from pyro_main import main_group_send_menu, main_create, send_new_msg, find_first_empty, add_new_category, \
    add_new_channel, delete_channel, delete_category, main_group_delete_menu, find_cell_by_link, ids, find_cell_by_id

admins = ['5582299570', '391275835', '763020856']
proc = False
api_id = 12498116
api_hash = "4e18f9670b086f276529be52f7f7f1a9"
app_bot = Client("my_bot", api_id=api_id, api_hash=api_hash, bot_token='5484106449:AAE7dylp6740RKXl4tXYjiap-99W7hmTt88')
app_user = Client("my_account", api_id=api_id, api_hash=api_hash)
file_ = openpyxl.load_workbook('list.xlsx')
sheet_obj_ = file_.active
i = 0
data_ = [({
    'channel': sheet_obj_.cell(row=i + 1, column=1).value,
    'category': sheet_obj_.cell(row=i + 1, column=2).value,
    'name_new': sheet_obj_.cell(row=i + 1, column=4).value,
    'id': sheet_obj_.cell(row=i + 1, column=5).value,
}) for i in range(1, find_first_empty(sheet_obj_)) if sheet_obj_.cell(row=i + 1, column=5).value]
channel_ids = [channel_id.get('channel').replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_') for
               channel_id in data_]

file_.close()


async def join_and_write(CallbackQuery):
    global proc
    proc = True
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active
    channels = [sheet_obj_c.cell(row=i + 1, column=1).value for i in range(1, find_first_empty(sheet_obj_c) - 1)]
    for k in range(len(channels)):
        print(channels[k])
        if not sheet_obj_c.cell(find_cell_by_link(channels[k], sheet_obj_c), 6).value:
            try:
                if "+" in channels[k]:
                    channels.append(channels[k])
                    row = find_cell_by_link(channels[k], sheet_obj_c)

                    join = await app_user.join_chat(channels[k])
                    print('join id '+str(join.id))
                    sheet_obj_c.cell(row, 6, value=join.id)
                    file_c.save('list.xlsx')
                else:
                    print(channels[k].replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_'))
                    name = channels[k].replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_')
                    row = find_cell_by_link(channels[k], sheet_obj_c)
                    join = await app_user.join_chat(name)
                    print('join id '+str(join.id))
                    sheet_obj_c.cell(row, 6, value=join.id)
                    file_c.save('list.xlsx')
            except FloodWait as ew:
                print('wait', ew.value)
                await asyncio.sleep(ew.value + 2)
                if "+" in channels[k]:
                    print(channels[k])
                    row = find_cell_by_link(channels[k], sheet_obj_c)
                    join = await app_user.join_chat(channels[k])
                    print('join id '+str(join.id))
                    sheet_obj_c.cell(row, 6, value=join.id)
                    file_c.save('list.xlsx')
                else:
                    print(channels[k].replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_'))
                    name = channels[k].replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_')
                    row = find_cell_by_link(channels[k], sheet_obj_c)
                    join = await app_user.join_chat(name)
                    print('join id '+str(join.id))
                    sheet_obj_c.cell(row, 6, value=join.id)
                    file_c.save('list.xlsx')
            except Exception as e:
                print("exception:",e)
            await asyncio.sleep(30)
    file_c.close()
    await main_create(app_bot, app_user, CallbackQuery)
    proc = False
    return channels


def return_categories():
    categories = []
    file_categories = openpyxl.load_workbook('categories.xlsx')
    sheet_obj_categories = file_categories.active

    for i in range(find_first_empty(sheet_obj_categories) - 1):
        if sheet_obj_categories.cell(row=i + 1, column=1).value not in categories:
            categories.append(sheet_obj_categories.cell(row=i + 1, column=1).value)

    file_categories.close()
    return categories


def return_channels():
    global i
    limit = i + 5
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active
    channels = []
    data_c = [({
        'channel': sheet_obj_c.cell(row=i + 1, column=1).value,
        'category': sheet_obj_c.cell(row=i + 1, column=2).value,
        'name_new': sheet_obj_c.cell(row=i + 1, column=4).value,
        'id': sheet_obj_c.cell(row=i + 1, column=5).value,
    }) for i in range(1, find_first_empty(sheet_obj_c) - 1)]
    while i < limit and i < len(data_c):
        channels.append(data_c[i].get('channel'))
        i += 1
    file_c.close()
    return channels


def return_all_channels():
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active

    data_c = [({
        'channel': sheet_obj_c.cell(row=j + 1, column=1).value,
        'category': sheet_obj_c.cell(row=j + 1, column=2).value,
        'name_new': sheet_obj_c.cell(row=j + 1, column=4).value,
        'id': sheet_obj_c.cell(row=i + 1, column=5).value,
    }) for j in range(1, find_first_empty(sheet_obj_c) - 1)]
    channel_ids = [channel_id.get('channel').replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_') for
                   channel_id in data_c]
    file_c.close()
    return channel_ids


def return_channel_btns(Btns):
    for c in return_channels():
        Btns.append([InlineKeyboardButton(c + '| Видалити', callback_data='//de_ch_app//' + c)])
    Btns.append([InlineKeyboardButton('Головне меню', callback_data='//main_menu//')])
    print(len(Btns))
    return Btns


def return_channels_ids():
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active

    data_c = [({
        'channel': sheet_obj_c.cell(row=j + 1, column=1).value,
        'category': sheet_obj_c.cell(row=j + 1, column=2).value,
        'name_new': sheet_obj_c.cell(row=j + 1, column=4).value,
        'id': sheet_obj_c.cell(row=i + 1, column=5).value,
    }) for j in range(1, find_first_empty(sheet_obj_c) - 1) if sheet_obj_.cell(row=j + 1, column=5).value]
    channel_ids = [channel_id.get('channel').replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_') for
                   channel_id in data_c]
    file_c.close()
    return channel_ids


async def send_bot_menu(app_bot, callback):
    MainButtons = [[InlineKeyboardButton('Головний канал', callback_data='//main_channel//'),
                    InlineKeyboardButton('Категорії', callback_data='//categories//')],
                   [InlineKeyboardButton('Канали', callback_data='//channels//')],
                   ]
    MainMarkup = InlineKeyboardMarkup(MainButtons)
    await callback.edit_message_text('**Головне меню**', reply_markup=MainMarkup)


async def send_bot_main_channel(app_bot, app_user, callback):
    MainButtons = [[InlineKeyboardButton('Надіслати меню', callback_data='//send_menu//'),
                    # InlineKeyboardButton('Видалити меню', callback_data='//delete_menu//')
                    ],
                   [InlineKeyboardButton('Головне меню', callback_data='//main_menu//')]
                   ]
    MainMarkup = InlineKeyboardMarkup(MainButtons)
    await callback.edit_message_text('**Головний канал**', reply_markup=MainMarkup)


async def send_bot_channels(app_bot, app_user, callback):
    MainButtons = [[InlineKeyboardButton('Створити канали', callback_data='//create_channels//')],
                   [InlineKeyboardButton('Додати канал', callback_data='//add_channel//'),
                    InlineKeyboardButton('Видалити канал', callback_data='//delete_channel//')],
                   [InlineKeyboardButton('Головне меню', callback_data='//main_menu//')]
                   ]
    MainMarkup = InlineKeyboardMarkup(MainButtons)
    await callback.edit_message_text('**Канали**', reply_markup=MainMarkup)


async def send_bot_categories(app_bot, app_user, callback):
    MainButtons = [
        [InlineKeyboardButton('Додати категорію', callback_data='//add_category//'),
         InlineKeyboardButton('Видалити категорію', callback_data='//delete_category//')],
        [InlineKeyboardButton('Головне меню', callback_data='//main_menu//')]
    ]
    MainMarkup = InlineKeyboardMarkup(MainButtons)
    await callback.edit_message_text('**Категорії**', reply_markup=MainMarkup)


@app_bot.on_message(filters.command('start') & filters.private & filters.create(
    lambda self, c, m: (str(m.from_user.id) in admins)))
async def on_start(app_bot, message):
    MainButtons = [[InlineKeyboardButton('Головний канал', callback_data='//main_channel//'),
                    InlineKeyboardButton('Категорії', callback_data='//categories//')],
                   [InlineKeyboardButton('Канали', callback_data='//channels//')],
                   ]
    MainMarkup = InlineKeyboardMarkup(MainButtons)
    await app_bot.send_message(message.chat.id, '**Головне меню**', reply_markup=MainMarkup)


@app_bot.on_message(filters.command('add_category') & filters.private & filters.create(
    lambda self, c, m: (str(m.from_user.id) in admins)))
async def add_category(app_bot, message):
    await add_new_category(app_bot, message)


@app_bot.on_message(filters.command('add_channel') & filters.private & filters.create(
    lambda self, c, m: (str(m.from_user.id) in admins)))
async def add_category(app_bot, message):
    await add_new_channel(app_bot, message)


@app_user.on_message(filters.channel and filters.create(lambda self, c, m: (m.chat.id in ids())))
async def check_updates(app_bot, message):
    new_channel_id = 0
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active
    for channel in ids():
        row = find_cell_by_id(channel, sheet_obj_c)
        if channel == message.chat.id:
            new_channel_id = sheet_obj_c.cell(row, 5).value
    file_c.close()

    await send_new_msg(app_bot, message, new_channel_id)


@app_bot.on_callback_query()
async def callback_query(app_bot, CallbackQuery):
    if CallbackQuery.data == '//main_menu//':
        await send_bot_menu(app_bot, CallbackQuery)
    elif CallbackQuery.data == '//categories//':
        await send_bot_categories(app_bot, app_user, CallbackQuery)
    elif CallbackQuery.data == '//channels//':
        await send_bot_channels(app_bot, app_user, CallbackQuery)
    elif CallbackQuery.data == '//main_channel//':
        await send_bot_main_channel(app_bot, app_user, CallbackQuery)
    elif CallbackQuery.data == '//send_menu//':
        await main_group_send_menu(app_bot, app_user, CallbackQuery)
    # elif CallbackQuery.data == '//delete_menu//':
    #     await main_group_delete_menu(app_user, CallbackQuery)
    elif CallbackQuery.data == '//add_category//':
        await app_bot.send_message(CallbackQuery.message.chat.id, '**Надрукуйте `/add_category ваша_назва_категорії`**')
        await CallbackQuery.answer()
    elif CallbackQuery.data == '//delete_category//':
        Btns = []
        for c in return_categories():
            if c != 'Категорії':
                Btns.append([InlineKeyboardButton(c + '| Видалити', callback_data='//delete_c_a//' + c)])
        Btns.append([InlineKeyboardButton('Головне меню', callback_data='//main_menu//')])
        await CallbackQuery.edit_message_text('**Оберіть категорії**', reply_markup=InlineKeyboardMarkup(Btns))
    elif '//delete_c_a//' in CallbackQuery.data:
        Btns = [
            [InlineKeyboardButton('Головне меню', callback_data='//main_menu//')],
            [InlineKeyboardButton('Видалити',
                                  callback_data='//delete_c_ad//' + CallbackQuery.data.split('//')[-1])]
        ]
        await CallbackQuery.edit_message_text(
            '***Ви впевнені, що хочете видалити категорію? Разом з нею видаляться усі канали з якими вона зв\'язана***',
            reply_markup=InlineKeyboardMarkup(Btns))
    elif '//delete_c_ad//' in CallbackQuery.data:
        await delete_category(app_bot, CallbackQuery.data.split('//')[-1], CallbackQuery)
        await send_bot_menu(app_bot, CallbackQuery)
    elif CallbackQuery.data == '//add_channel//':
        await app_bot.send_message(CallbackQuery.message.chat.id,
                                   '**Надрукуйте `/add_channel ваш_парсинг_канал, ваша_категорія, ваша_нова_назва_каналу`**')
        await CallbackQuery.answer()
    elif CallbackQuery.data == '//delete_channel//':
        Btns = []
        global i
        i = 0
        Btns = return_channel_btns(Btns)
        Btns.append([InlineKeyboardButton('Наступна...', callback_data='//add5//')])

        await CallbackQuery.edit_message_text('**Оберіть канал**', reply_markup=InlineKeyboardMarkup(Btns))
    elif '//add5//' in CallbackQuery.data:
        Btns = []
        Btns = return_channel_btns(Btns)
        if i < len(return_all_channels()):
            Btns.append([InlineKeyboardButton('Наступна...', callback_data='//add5//')])
        await CallbackQuery.edit_message_text('**Оберіть канал**', reply_markup=InlineKeyboardMarkup(Btns))
    elif '//de_ch_app//' in CallbackQuery.data:
        await delete_channel(app_bot, CallbackQuery.data.replace('//delete_channel_approved//', '').strip(),
                             CallbackQuery)
        await send_bot_menu(app_bot, CallbackQuery)

    elif CallbackQuery.data == '//create_channels//':
        global proc
        if not proc:
            await join_and_write(CallbackQuery)
            await asyncio.sleep(2)


app_user.start()
app_bot.run()
