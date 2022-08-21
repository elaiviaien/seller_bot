import asyncio
import random
import time
from pyrogram.errors.exceptions.bad_request_400 import UsernameNotOccupied
from pyrogram.errors import UsernameInvalid, FloodWait, UsernameNotOccupied
from pyrogram.raw import functions
from slugify import slugify
import openpyxl
from pyrogram import Client, filters
from pyrogram.types import ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton

from pyro_main import main_group_send_menu, main_create, send_new_msg

api_id = 10736822
api_hash = "3a730347f1f410c6d8491fbfaed0add9"
app_bot = Client("my_bot", api_id=api_id, api_hash=api_hash, bot_token='5784096253:AAEnhY3_m2WLalGWStmkYD5R2DQixbFsmyo')
app_user = Client("my_account", api_id=api_id, api_hash=api_hash)
file_ = openpyxl.load_workbook('list.xlsx')
sheet_obj_ = file_.active

data_ = [({
    'channel': sheet_obj_.cell(row=i + 1, column=1).value,
    'category': sheet_obj_.cell(row=i + 1, column=2).value,
    'name_new': sheet_obj_.cell(row=i + 1, column=4).value,
    'id': sheet_obj_.cell(row=i + 1, column=5).value,
}) for i in range(1,sheet_obj_.max_row) if sheet_obj_.cell(row=i + 1, column=5).value]
channel_ids = [channel_id.get('channel').replace('https://t.me/', '').replace('joinchat/', '').replace('-', '_') for
               channel_id in data_]
file_.close()



app_bot.start()
app_user.run()
