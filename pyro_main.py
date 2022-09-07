import asyncio
import os
import random
from pyrogram.errors import FloodWait
import openpyxl
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from PIL import Image, ImageFilter


def ids():
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_c = file_c.active
    ids = [sheet_obj_c.cell(row=i + 1, column=6).value for i in range(1, find_first_empty(sheet_obj_c) - 1)]
    file_c.close()
    return ids


def find_first_empty(self):
    r = 0
    while True:
        r += 1
        if not self.cell(r, 1).value:
            return r


main_group_id = 'Turk_0pt'
file_ = openpyxl.load_workbook('list.xlsx')
sheet_obj_ = file_.active
categories = []
row_ = 1
file_categories = openpyxl.load_workbook('categories.xlsx')
sheet_obj_categories = file_categories.active

data_ = [({
    'channel': sheet_obj_.cell(row=i + 1, column=1).value,
    'category': sheet_obj_.cell(row=i + 1, column=2).value,
    'name_new': sheet_obj_.cell(row=i + 1, column=4).value,
    'id': sheet_obj_.cell(row=i + 1, column=5).value,
}) for i in range(1, find_first_empty(sheet_obj_) - 1)]
def return_categories():
    categories = []
    file_categories = openpyxl.load_workbook('categories.xlsx')
    sheet_obj_categories = file_categories.active

    for i in range(find_first_empty(sheet_obj_categories) - 1):
        if sheet_obj_categories.cell(row=i + 1, column=1).value not in categories:
            categories.append(sheet_obj_categories.cell(row=i + 1, column=1).value)

    file_categories.close()
    return categories
def return_data_():
    file_ = openpyxl.load_workbook('list.xlsx')
    sheet_obj_ = file_.active
    file_categories = openpyxl.load_workbook('categories.xlsx')
    data_ = [({
        'channel': sheet_obj_.cell(row=i + 1, column=1).value,
        'category': sheet_obj_.cell(row=i + 1, column=2).value,
        'name_new': sheet_obj_.cell(row=i + 1, column=4).value,
        'id': sheet_obj_.cell(row=i + 1, column=5).value,
    }) for i in range(1, find_first_empty(sheet_obj_) - 1)]
    file_categories.close()
    file_.close()
    return data_


def find_cell_by_link(v, sheet_obj):
    i = 0
    while True:
        i += 1
        if sheet_obj.cell(i, 1).value == v:
            return i


def find_cell_by_id(v, sheet_obj):
    i = 0
    while True:
        i += 1
        if sheet_obj.cell(i, 6).value == v:
            return i


async def main_create(app_bot, app_user, CallbackQuery):
    ids_ = ids()
    for i in range(len(return_data_())):
        print('iter')
        await asyncio.sleep(10)
        await create_channels(return_data_()[i].get('channel'), app_bot, app_user, CallbackQuery, ids_[i])
    await app_bot.send_message(CallbackQuery.message.chat.id,
                               'Канали створені')


async def create_channel(channel, app_user, chat_id):
    await app_user.promote_chat_member(channel.id, "Turk_0ptbot")
    last_mes = None
    mg_id = 0
    async for message in app_user.get_chat_history(
            chat_id, limit=10000):
        if not last_mes:
            last_mes = message
        if not mg_id and not message.media_group_id:
            message.media_group_id = 0

        if (message.photo or message.video or last_mes.photo or last_mes.media_group_id) and mg_id != message.media_group_id:
            last_mes = message
            if message.media_group_id:
                try:
                    await app_user.copy_media_group(channel.id,
                                                    message.chat.id, message.id)
                    kk = random.randrange(10, 30)
                    await asyncio.sleep(kk)
                except FloodWait as e:
                    print('wait for', e.value, 'to send message')
                    await asyncio.sleep(e.value + 2)
                    await app_user.copy_media_group(channel.id,
                                                    message.chat.id, message.id)
                    kk = random.randrange(10, 30)
                    await asyncio.sleep(kk)
                except Exception as e:
                    print(e)
            elif message.photo:

                try:
                    await app_user.send_photo(channel.id, message.photo.file_id)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except FloodWait as e:
                    print('wait for', e.value, 'to send message')
                    await asyncio.sleep(e.value + 2)
                    await app_user.send_photo(channel.id, message.photo.file_id)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except Exception as e:
                    print(e)
            elif message.video:
                try:
                    await app_user.send_video(channel.id, message.video.file_id)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except FloodWait as e:
                    print('wait for', e.value, 'to send message')
                    await asyncio.sleep(e.value + 2)
                    await app_user.send_video(channel.id, message.video.file_id)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except Exception as e:
                    print(e)
            if message.text:
                try:
                    await app_user.send_message(channel.id, message.text)
                    print('text')
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except FloodWait as e:
                    print('wait for', e.value, 'to send message')
                    await asyncio.sleep(e.value + 2)
                    await app_user.send_message(channel.id, message.text)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except Exception as e:
                    print(e)
            if message.caption:
                try:
                    await app_user.send_message(channel.id, message.caption)
                    print('text')
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except FloodWait as e:
                    print('wait for', e.value, 'to send message')
                    await asyncio.sleep(e.value + 2)
                    await app_user.send_message(channel.id, message.caption)
                    kk = random.randrange(1, 10)
                    await asyncio.sleep(kk)
                except Exception as e:
                    print(e)
        mg_id = message.media_group_id


async def create_channels(title, app_bot, app_user, CallbackQuery, chat_id):
    file = openpyxl.load_workbook('list.xlsx')
    sheet_obj = file.active
    row = find_cell_by_link(title, sheet_obj)
    if sheet_obj.cell(row=row, column=5).value:
        try:
            print(sheet_obj.cell(row=row, column=5).value)
            msg = await app_user.get_messages(chat_id=sheet_obj.cell(row=row, column=5).value, message_ids=1)
        except Exception as e:
            print(e)
        file.close()
        file.close()
        await app_bot.send_message(CallbackQuery.message.chat.id,
                                   'Канал вже існує: ' + str(sheet_obj.cell(row=row, column=4).value))
        print('channel exists')
    else:
        try:
            channel = await app_user.create_channel(sheet_obj.cell(row=row, column=4).value)
            sheet_obj.cell(row=row, column=5, value=channel.id)
            file.save('list.xlsx')
            file.close()
            await asyncio.sleep(3)
            await create_channel(channel, app_user, chat_id)
        except FloodWait as e:
            print('wait for', e.value, "to create channel")
            await app_bot.send_message(CallbackQuery.message.chat.id,
                                       'Почекайте ' + str(e.value) + " секунди щоб створити канал")
            await asyncio.sleep(e.value + 2)
            channel = await app_user.create_channel(sheet_obj.cell(row=row, column=4).value)
            sheet_obj.cell(row=row, column=5, value=channel.id)
            file.save('list.xlsx')
            file.close()
            await asyncio.sleep(3)
            await create_channel(channel, app_user, chat_id)


async def main_group_send_menu(app_bot, app_user, CallbackQuery):
    global main_group_id
    CategoriesButtons = []
    print(return_categories())
    for c in return_categories():
        GroupsButtons = []
        cc = 0
        for g in [el for el in return_data_() if el.get('category') == c]:
            if g.get('id') and cc<9:
                cc+=1
                url = await app_user.create_chat_invite_link(g.get('id'))
                btn_g = InlineKeyboardButton(g.get('name_new'), url=url.invite_link, )
                GroupsButtons.append([btn_g])

        await asyncio.sleep(1)
        if GroupsButtons:
            msg = await app_bot.send_message(main_group_id, c, reply_markup=InlineKeyboardMarkup(GroupsButtons))
            file = openpyxl.load_workbook('categories.xlsx')
            sheet_obj = file.active
            row = find_cell_by_link(c, sheet_obj)
            sheet_obj.cell(row=row, column=2, value=msg.link)
            file.save('categories.xlsx')
            file.close()
            print(sheet_obj.cell(row=row, column=2).value)
            btn = InlineKeyboardButton(c, url=msg.link)
            CategoriesButtons.append([btn])
        cc_n= 0
        GroupsButtons = []
        for g in [el for el in return_data_() if el.get('category') == c]:
            cc_n+=1
            if g.get('id') and 9 <= cc_n:
                url = await app_user.create_chat_invite_link(g.get('id'))
                btn_g = InlineKeyboardButton(g.get('name_new'), url=url.invite_link, )
                GroupsButtons.append([btn_g])
        await asyncio.sleep(1)
        if GroupsButtons:
            msg = await app_bot.send_message(main_group_id, c, reply_markup=InlineKeyboardMarkup(GroupsButtons))
    CategoriesMarkup = InlineKeyboardMarkup(
        CategoriesButtons
    )





    cat_m = await app_bot.send_message(main_group_id, 'Категорії', reply_markup=CategoriesMarkup)
    file = openpyxl.load_workbook('categories.xlsx')
    sheet_obj = file.active
    row = find_cell_by_link('Категорії', sheet_obj)
    sheet_obj.cell(row=row, column=2, value=cat_m.link)
    file.save('categories.xlsx')
    file.close()
    await app_bot.send_message(CallbackQuery.message.chat.id,
                               'Меню надіслане')


async def send_new_msg(app_user, message, new_channel_id):
    last_mes = await app_user.get_messages(message.chat.id, message.id - 1)
    mg_id = await app_user.get_messages(message.chat.id, message.id - 1)
    mg_id = mg_id.media_group_id
    if not mg_id and not message.media_group_id:
        message.media_group_id = 0

    if (message.photo or message.video or last_mes.photo or last_mes.media_group_id) and mg_id != message.media_group_id:
        kk = random.randrange(1, 4)
        await asyncio.sleep(kk)
        if message.media_group_id:

            try:
                await app_user.copy_media_group(new_channel_id,
                                                message.chat.id, message.id)
            except FloodWait as e:
                print('wait for', e.value, 'to send message')
                await asyncio.sleep(e.value + 2)
                await app_user.copy_media_group(new_channel_id,
                                                message.chat.id, message.id)
            except Exception as e:
                print(e)
        elif message.photo:

            try:
                await app_user.send_photo(new_channel_id, message.photo.file_id)
            except FloodWait as e:
                print('wait for', e.value, 'to send message')
                await asyncio.sleep(e.value + 2)
                await app_user.send_photo(new_channel_id, message.photo.file_id)
            except Exception as e:
                print(e)
        elif message.video:
            try:
                await app_user.send_video(new_channel_id, message.video.file_id)
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except FloodWait as e:
                print('wait for', e.value, 'to send message')
                await asyncio.sleep(e.value + 2)
                await app_user.send_video(new_channel_id, message.video.file_id)
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except Exception as e:
                print(e)
        if message.text:
            try:
                await app_user.send_message(new_channel_id, message.text)
                print('text')
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except FloodWait as e:
                print('wait for', e.value, 'to send message')
                await asyncio.sleep(e.value + 2)
                await app_user.send_message(new_channel_id, message.text)
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except Exception as e:
                print(e)
        if message.caption:
            try:
                await app_user.send_message(new_channel_id, message.caption)
                print('text')
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except FloodWait as e:
                print('wait for', e.value, 'to send message')
                await asyncio.sleep(e.value + 2)
                await app_user.send_message(new_channel_id, message.caption)
                kk = random.randrange(1, 10)
                await asyncio.sleep(kk)
            except Exception as e:
                print(e)


async def add_new_category(app_bot, message):
    file_c = openpyxl.load_workbook('categories.xlsx')
    sheet_obj_n = file_c.active
    row = find_first_empty(sheet_obj_n)
    name = message.text.replace('/add_category', '').strip()
    categories = []
    for i in range(find_first_empty(sheet_obj_n)):
        if sheet_obj_n.cell(row=i + 1, column=1).value not in categories:
            categories.append(sheet_obj_n.cell(row=i + 1, column=1).value)
    if name not in categories:
        sheet_obj_n.cell(row, 1, value=name)
        file_c.save('categories.xlsx')
        file_c.close()
        await app_bot.send_message(message.chat.id, 'Категорія додана')
    else:
        await app_bot.send_message(message.chat.id, 'Така категорія вже існує')
    await asyncio.sleep(3)


async def add_new_channel(app_bot, message):
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_n = file_c.active
    row = find_first_empty(sheet_obj_n)
    name = message.text.replace('/add_channel', '')
    channels = []
    for i in range(1, find_first_empty(sheet_obj_n)):
        if sheet_obj_n.cell(row=i + 1, column=1).value not in channels:
            channels.append(sheet_obj_n.cell(row=i + 1, column=1).value)
    name = name.split(',')
    if name[0] not in channels:
        try:
            sheet_obj_n.cell(row, 1, value=name[0].strip())
            sheet_obj_n.cell(row, 2, value=name[1].strip())
            sheet_obj_n.cell(row, 4, value=name[2].strip())
            await app_bot.send_message(message.chat.id, 'Канал додано')

            file_c.save('list.xlsx')
        except Exception as e:
            print(e)
            await app_bot.send_message(message.chat.id, 'Неправильно введені дані')


    else:
        await app_bot.send_message(message.chat.id, 'Такий канал вже додано')
    file_c.close()
    await asyncio.sleep(3)


async def delete_channel(app_bot, message, CallbackQuery):
    file_c = openpyxl.load_workbook('list.xlsx')
    sheet_obj_n = file_c.active
    # row = find_first_empty(sheet_obj_n)
    name = message.replace('//de_ch_app//', '')
    channels = []
    print(name)
    for i in range(1, find_first_empty(sheet_obj_n)):
        if sheet_obj_n.cell(row=i + 1, column=1).value not in channels:
            channels.append(sheet_obj_n.cell(row=i + 1, column=1).value)
    name = name.replace(' ', '')
    if name in channels:
        try:
            row = find_cell_by_link(name, sheet_obj_n)
            sheet_obj_n.delete_rows(row, 1)
            await app_bot.send_message(CallbackQuery.message.chat.id, 'Канал видалено')
        except Exception as e:
            print(e)
            await app_bot.send_message(CallbackQuery.message.chat.id, 'Помилка при видаленні')
        file_c.save('list.xlsx')
        file_c.close()
    else:
        await app_bot.send_message(CallbackQuery.message.chat.id, 'Немає такого каналу')
    await CallbackQuery.answer()

    await asyncio.sleep(3)


async def delete_category(app_bot, message, CallbackQuery):
    file_c = openpyxl.load_workbook('categories.xlsx')
    sheet_obj_n = file_c.active
    # row = find_first_empty(sheet_obj_n)

    name = message.replace('/delete_category', '').strip()
    categories = []
    for i in range(find_first_empty(sheet_obj_n)):
        if sheet_obj_n.cell(row=i + 1, column=1).value not in categories:
            categories.append(sheet_obj_n.cell(row=i + 1, column=1).value)
    print(name)
    deleted = False
    if name in categories:
        try:
            row = find_cell_by_link(name, sheet_obj_n)
            sheet_obj_n.delete_rows(row, 1)
            deleted = True
            await app_bot.send_message(CallbackQuery.message.chat.id, 'Категорія видалена')
        except Exception as e:
            print(e)
            await app_bot.send_message(CallbackQuery.message.chat.id, 'Помилка при видаленні')
        file_c.save('categories.xlsx')
        file_c.close()

        print(deleted)
        if deleted:
            file_c = openpyxl.load_workbook('list.xlsx')
            sheet_obj_n = file_c.active
            for i in range(1, find_first_empty(sheet_obj_n)):
                if sheet_obj_n.cell(i, 2).value == name:
                    sheet_obj_n.delete_rows(i, 1)
                    file_c.save('list.xlsx')
            file_c.close()
    else:
        await app_bot.send_message(CallbackQuery.message.chat.id, 'Немає такої категорії')
    await asyncio.sleep(3)


async def main_group_delete_menu(app_user, CallbackQuery):
    file_c = openpyxl.load_workbook('categories.xlsx')
    sheet_obj_n = file_c.active
    menu_msgs = []
    for i in range(1, find_first_empty(sheet_obj_n)):
        if sheet_obj_n.cell(i, 2).value:
            menu_msgs.append(int(sheet_obj_n.cell(i, 2).value.split('/')[-1]))
    await app_user.delete_messages(main_group_id, menu_msgs)
    await asyncio.sleep(3)
